import os, time
import streamlit as st
from pathlib import Path
import unicodedata

import base64
from pathlib import Path
import streamlit as st



def _logo_b64(path: str = "assets/company_logo.png") -> str:
    """Retourne le logo encodé en base64 (cherche à plusieurs emplacements)."""
    candidates = [
        Path(path),
        Path.cwd() / path,
        Path.cwd() / "assets" / "company_logo.png",
        Path(__file__).resolve().parent / path,
        Path(__file__).resolve().parent / "assets" / "company_logo.png",
    ]
    for p in candidates:
        try:
            if p.exists():
                return base64.b64encode(p.read_bytes()).decode("utf-8")
        except Exception:
            pass
    return ""

def inject_brand_css():
    brand_blue   = "#0C3D91"
    brand_orange = "#F7941D"
    light_text   = "#FFFFFF"
    dark_text    = "#0B1F44"
    bg = "#042B80"
    pattern_opacity = 0.012

    logo_b64 = _logo_b64()

    st.markdown(f"""
    <style>
      /* ===== Fond + filigrane ===== */
      .stApp {{
        background:
          radial-gradient(rgba(7,28,71,{pattern_opacity}) 1px, transparent 1px) 0 0/10px 10px,
          linear-gradient(160deg, {bg} 0%, {bg} 45%, {bg} 100%);
        background-attachment: fixed;
      }}
      {f'.stApp::before {{ content:""; position:fixed; inset:0; background:url("data:image/png;base64,{logo_b64}") no-repeat 24px 24px; background-size:160px; opacity:.12; pointer-events:none; z-index:0; }}' if logo_b64 else ''}

      /* ===== Titres & labels en BLANC (fond sombre) ===== */
      .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {{ color:{light_text} !important; }}
      .stApp .stSelectbox > label, .stApp .stMultiSelect > label,
      .stApp .stTextInput > label, .stApp .stNumberInput > label,
      .stApp .stDateInput > label, .stApp .stTextArea > label,
      .stApp .stSlider > label, .stApp .stRadio > label,
      .stApp .stCheckbox > label, .stApp label {{ color:{light_text} !important; }}
      .stApp .stMarkdown, .stApp .markdown-text-container {{ color:{light_text} !important; }}

      /* ===== Sélecteurs ===== */
      .stApp div[data-baseweb="select"], .stApp div[data-baseweb="select"] * {{ color:#111 !important; fill:#111 !important; }}
      .stApp [data-baseweb="select"] input::placeholder {{ color:rgba(0,0,0,.55) !important; }}
      body [data-baseweb="layer"] [role="listbox"],
      body [data-baseweb="popover"] [role="listbox"] {{
        background:#FFF !important; border:1px solid rgba(12,61,145,.35) !important; box-shadow:0 8px 24px rgba(7,28,71,.18);
      }}
      body [role="listbox"] [role="option"], body [role="listbox"] [role="option"] * {{ color:#111 !important; fill:#111 !important; opacity:1 !important; }}
      body [role="listbox"] [role="option"]:hover, body [role="listbox"] [role="option"]:hover * {{
        background:#F3F6FB !important; color:#111 !important; fill:#111 !important;
      }}
      body [role="listbox"] [role="option"][aria-selected="true"],
      body [role="listbox"] [role="option"][aria-selected="true"] * {{
        background:#FFE8E8 !important; color:#B21F2D !important; fill:#B21F2D !important;
      }}

      /* ===== Chips indisponibilités ===== */
      [data-baseweb="tag"] {{ background:#E9F4FF !important; border:1px solid rgba(12,61,145,.35) !important; }}
      [data-baseweb="tag"] * {{ color:{dark_text} !important; }}
      [data-baseweb="tag"] svg {{ fill:{brand_blue} !important; }}
      .unavail [data-baseweb="tag"] {{ background:rgba(220,53,69,.12) !important; border:1px solid rgba(220,53,69,.60) !important; }}
      .unavail [data-baseweb="tag"] *, .unavail [data-baseweb="tag"] svg {{ color:#7a0c0c !important; fill:#7a0c0c !important; }}

      /* ===== Boutons / Onglets ===== */
      div[data-baseweb="tab-list"], div[role="tablist"] {{ gap:12px !important; border-bottom:none !important; padding-bottom:8px; }}
      div[data-baseweb="tab-list"] button, div[role="tablist"] > button[role="tab"] {{
        background:#FFF !important; border:1px solid rgba(12,61,145,.18) !important;
        border-radius:999px !important; padding:.45rem .9rem !important;
        box-shadow:0 1px 1px rgba(7,28,71,.06); font-weight:600 !important; color:#000 !important;
      }}
      div[data-baseweb="tab-list"] button[aria-selected="true"],
      div[role="tablist"] > button[role="tab"][aria-selected="true"] {{ border-color:{brand_orange} !important; box-shadow:0 2px 6px rgba(247,148,29,.25); }}
      div[data-baseweb="tab-highlight"], div[role="tablist"] > div[aria-hidden="true"] {{ background:{brand_orange} !important; height:3px !important; border-radius:2px; }}

      .stButton>button {{ background:{brand_orange}; color:#fff; border:0; border-radius:10px; padding:.55rem 1rem; box-shadow:0 3px 0 #d17f12; }}
      .stButton>button:hover {{ background:#FFA23A; }}
      .stApp [data-testid="stFormSubmitButton"] button,
      .stApp [data-testid="stForm"] button,
      .stApp form button,
      .stApp button[kind][data-testid^="baseButton"] {{
        background:{brand_orange} !important; color:#fff !important; border:0 !important; border-radius:10px !important;
        padding:.55rem 1rem !important; box-shadow:0 3px 0 #d17f12 !important;
      }}
      .stApp [data-testid="stFormSubmitButton"] button:hover,
      .stApp [data-testid="stForm"] button:hover,
      .stApp form button:hover,
      .stApp button[kind][data-testid^="baseButton"]:hover {{ background:#FFA23A !important; }}

      /* ===== HÉRO & TITRE ===== */
      .welcome-wrap {{ display:flex; justify-content:center; margin: 18px 0 10px; }}
      .welcome-card {{
        background:{brand_orange}; color:#fff; padding:22px 28px; border-radius:16px;
        box-shadow:0 10px 24px rgba(0,0,0,.18);
        max-width:880px; width:min(92vw,880px); text-align:center;
      }}
      .welcome-card h2 {{ margin:0 0 6px 0; font-weight:800; font-size:clamp(22px, 3.2vw, 34px); }}
      .welcome-card p  {{ margin:0; opacity:.95; font-size:clamp(12px, 1.4vw, 16px); }}
      .page-title {{ text-align:center; margin: 8px 0 14px; font-size: clamp(26px, 4vw, 44px); }}

      /* ===== Cartes d’alerte (surfaces blanches) ===== */
      .stApp .stAlert {{
        background:#fff !important;
        border:1px solid rgba(12,61,145,.25) !important;
        border-radius:10px !important;
        box-shadow:0 6px 18px rgba(7,28,71,.08);
        color:{dark_text} !important;
      }}
      .stApp .stAlert * {{ color:{dark_text} !important; }}

      /* ===== Notice blanche + texte rouge ===== */
      .stApp [data-testid="stMarkdownContainer"] .notice-white-red,
      .stApp .stMarkdown .notice-white-red,
      .stApp .markdown-text-container .notice-white-red {{
        background:#fff !important;
        border:2px solid #dc2626 !important;
        border-radius:10px;
        padding:.75rem 1rem;
        box-shadow:0 6px 18px rgba(7,28,71,.10);
        color:#dc2626 !important;
        display:block;
        margin:10px 0 24px !important;
      }}
      .stApp [data-testid="stMarkdownContainer"] .notice-white-red *,
      .stApp .stMarkdown .notice-white-red *,
      .stApp .markdown-text-container .notice-white-red * {{ color:#dc2626 !important; }}

      /* ===== Sidebar – forcer le texte (titres/markdown) en NOIR ===== */
      .stApp [data-testid="stSidebar"] .stMarkdown,
      .stApp [data-testid="stSidebar"] .markdown-text-container,
      .stApp [data-testid="stSidebar"] h1,
      .stApp [data-testid="stSidebar"] h2,
      .stApp [data-testid="stSidebar"] h3,
      .stApp [data-testid="stSidebar"] h4,
      .stApp [data-testid="stSidebar"] h5,
      .stApp [data-testid="stSidebar"] h6,
      .stApp [data-testid="stSidebar"] label {{
        color:#111 !important;
      }}

      /* ===== Utilitaires ===== */
      .force-black, .force-black * {{ color:#111 !important; }}  /* -> pour forcer le noir dans un bloc précis */

    </style>
    """, unsafe_allow_html=True)



def alert_white_red(msg: str):
    """Bloc fond blanc + texte rouge (protégé contre l'héritage global)."""
    st.markdown(f"<div class='notice-white-red'>{msg}</div>", unsafe_allow_html=True)



def unavail_multiselect(label, options, key=None, **kwargs):
    """Multiselect avec chips rouges pour indisponibilités."""
    st.markdown('<div class="unavail">', unsafe_allow_html=True)
    value = st.multiselect(label, options, key=key, **kwargs)
    st.markdown('</div>', unsafe_allow_html=True)
    return value



inject_brand_css()

# =================== Auth ===================
def _logout():
    """Nettoie la session et relance l'app."""
    st.session_state.pop("auth", None)
    st.session_state.pop("last_activity", None)
    try:
        st.cache_data.clear()
        st.cache_resource.clear()
    except Exception:
        pass
    st.rerun()

def require_login(idle_timeout_min: int | None = None):
    """
    Protège l'app par mot de passe.
    - Le mot de passe est lu dans st.secrets['app']['password'] ou APP_PASSWORD.
    - idle_timeout_min : déconnexion auto après X minutes d'inactivité (None pour désactiver).
    """
    PASSWORD = st.secrets.get("app", {}).get("password") or os.environ.get("APP_PASSWORD")

    # Déjà authentifié ?
    if st.session_state.get("auth"):
        # Déconnexion auto si inactif
        if idle_timeout_min is not None:
            now = time.time()
            last = st.session_state.get("last_activity", now)
            if (now - last) > idle_timeout_min * 60:
                _logout()
            else:
                st.session_state["last_activity"] = now

        # Bouton "Se déconnecter" (sidebar)
        if st.sidebar.button("🔓 Se déconnecter", use_container_width=True):
            _logout()
        return  # laisser le reste de l'app s'exécuter

    # Écran de connexion (si non authentifié)
    st.title("🔐 Accès protégé")


    with st.form("login_form"):
        pwd = st.text_input("Mot de passe", type="password")
        ok = st.form_submit_button("Entrer")
    if ok:
        if PASSWORD and pwd == PASSWORD:
            st.session_state.auth = True
            st.session_state.last_activity = time.time()
            st.rerun()  # recharge l’UI immédiatement
        else:
            st.error("Mot de passe incorrect")
    st.stop()
# ============================================

# 👉 Appelle l’auth AVANT d’afficher l’UI
require_login(idle_timeout_min=60)  # mets None pour désactiver le timeout




    
import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from optimizer import (
    run_optimization,
    update_kilometrage_file,
    update_distance_parcourue
)
import folium
from streamlit_folium import folium_static

# I/O Google Drive
from drive_io import drive_download, drive_upload

# ===================== PDF (ReportLab) =====================
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
except Exception as _e:
    _PDF_IMPORT_ERROR = _e
else:
    _PDF_IMPORT_ERROR = None

import re
import numpy as np

def _deferred_reset(flag_key: str, keys: list[str]):
    """Si flag actif, supprime les clés de widgets AVANT instanciation, puis éteint le flag."""
    if st.session_state.get(flag_key):
        for k in keys:
            st.session_state.pop(k, None)
        st.session_state[flag_key] = False

# ======== Commandes : normalisation & import local ========
REQUIRED_ORDER_COLS = [
    "Code client", "Libellé", "Quantité", "Unité", "Adresse", "Code postal", "Ville"
]

def _canon_col(name: str) -> str:
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().lower()

# Synonymes tolérés -> nom cible
_COL_SYNONYMS = {
    "Code client": {"code client","client","code","code tiers","code (tiers)","code_client","client code"},
    "Libellé": {"libelle","designation","produit","article","label","intitule"},
    "Quantité": {"quantite","qte","qty","quantite totale","quantite commandee","quantity"},
    "Unité": {"unite","unit","uom"},
    "Adresse": {"adresse","adresse 1","adresse complete","adresse complète","adress"},
    "Code postal": {"code postal","cp","postal code","zipcode","zip"},
    "Ville": {"ville","commune","city","localite","localité"},
}

def _standardize_orders_df(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    rename = {}
    for c in list(df.columns):
        cc = _canon_col(c)
        for target, syns in _COL_SYNONYMS.items():
            if cc in syns:
                rename[c] = target
                break
    df = df.rename(columns=rename)
    missing = [col for col in REQUIRED_ORDER_COLS if col not in df.columns]
    return df, missing

def _uploaded_orders_to_excelbuf(uploaded) -> tuple[BytesIO, pd.DataFrame, list[str]]:
    """Lit xlsx/csv, normalise colonnes, renvoie un buffer Excel prêt pour optimizer."""
    # Lire
    name = (uploaded.name or "").lower()
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded)
    elif name.endswith(".csv"):
        # détection auto du séparateur
        df = pd.read_csv(uploaded, sep=None, engine="python")
    else:
        raise ValueError("Format non supporté (utilise .xlsx / .xls / .csv).")

    # Normaliser colonnes
    df, missing = _standardize_orders_df(df)
    if missing:
        return None, df, missing

    # Sauver en Excel (in-memory) pour rester compatible avec run_optimization
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    out.seek(0)
    return out, df, []

# -------- Helpers nom chauffeur / parsing / métriques ----------
def _clean_chauffeur_name(name: str) -> str:
    """Supprime un éventuel préfixe 'Chauffeur X :' pour n’afficher que le vrai nom."""
    return re.sub(r"^\s*Chauffeur\s*\d+\s*:\s*", "", str(name)).strip()

def _extract_chauffeur_vehicle(header_line: str):
    """
    Extrait (chauffeur, véhicule) depuis une ligne type :
    'Tournée 1 : Chauffeur 1 : Soumaila via IVECO FK-233-XA'
    """
    m = re.search(r"Tournée\s*\d+\s*:\s*(.*?)\s+via\s+(.*)", header_line)
    if m:
        ch = _clean_chauffeur_name(m.group(1).strip())
        vh = m.group(2).strip()
        return ch, vh
    # fallback
    parts = header_line.split("via")
    ch = _clean_chauffeur_name(parts[0].split(":", 1)[-1].strip() if parts else "")
    vh = parts[1].strip() if len(parts) > 1 else ""
    return ch, vh

def _aggregate_orders_for_pdf(orders_df: pd.DataFrame) -> dict:
    """Retourne {code_client: {'Poids': kg, 'Cartons': nb}} pour alimenter le PDF."""
    df = orders_df.copy()
    for col in ["Code client", "Unité", "Quantité"]:
        if col not in df.columns:
            df[col] = np.nan
    df["Quantité"] = pd.to_numeric(df["Quantité"], errors="coerce").fillna(0.0)
    df["Poids"]    = np.where(df["Unité"].astype(str).str.upper().eq("KG"), df["Quantité"], 0.0)
    df["Cartons"]  = np.where(df["Unité"].astype(str).str.upper().eq("U"),  df["Quantité"] / 30.0, 0.0)
    agg = df.groupby("Code client", dropna=True).agg({"Poids":"sum","Cartons":"sum"})
    return agg.to_dict(orient="index")

def _generate_pdf_tours(assigned: list, df_geo: pd.DataFrame, orders_df: pd.DataFrame, result_text: str) -> BytesIO:
    """
    Génère un PDF multi-tournées.
    - 'Restaurants' = nombre de clients effectivement imprimés (nœuds != dépôt).
    - Totaux Poids/Cartons = sommes sur ces mêmes clients (depuis orders_df).
    """
    # --- Mapping géocodage ---
    code_col = next((c for c in df_geo.columns if "Code" in c and "tiers" in c), None)
    if not code_col:
        code_col = "Code (tiers)"
        if code_col not in df_geo.columns:
            df_geo[code_col] = "FRESH DISTRIB"
    geo_map = df_geo.set_index(code_col).to_dict(orient="index")

    # --- Poids/Cartons par client (depuis commandes) ---
    metrics = _aggregate_orders_for_pdf(orders_df)  # {code: {"Poids": kg, "Cartons": nb}}

    # --- Styles PDF ---
    styles = getSampleStyleSheet()
    SCALE = 0.9  # 90% de la taille actuelle

    def fs(x): return round(x * SCALE, 1)
    def ld(x): return round((x * SCALE) + 2, 1)  # leading un poil > fontSize

    style_title     = ParagraphStyle("title",     parent=styles["Heading1"],
                                    fontName="Helvetica-Bold", fontSize=fs(26), leading=ld(26),
                                    textColor=colors.black, alignment=1, spaceAfter=fs(6))
    style_hdr_right = ParagraphStyle("hdr_right", parent=styles["Normal"],
                                    fontName="Helvetica-Bold", fontSize=fs(12), leading=ld(12), alignment=2)
    style_line_big  = ParagraphStyle("line_big",  parent=styles["Normal"],
                                    fontName="Helvetica-Bold", fontSize=fs(12), leading=ld(12), spaceAfter=fs(4))
    style_line_small= ParagraphStyle("line_small",parent=styles["Normal"],
                                    fontName="Helvetica",      fontSize=fs(9),  leading=ld(9))
    style_client    = ParagraphStyle("client",    parent=styles["Normal"],
                                    fontName="Helvetica-Bold", fontSize=fs(11), leading=ld(11))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.0*cm, rightMargin=1.0*cm,
                            topMargin=0.8*cm, bottomMargin=0.8*cm)
    flow = []

    for k, tour in enumerate(assigned, start=1):
        tour_id   = tour["tour_id"]
        chauffeur = tour["chauffeur"]    # ex. "Soumaila"
        veh       = tour["vehicle"]      # ex. "IVECO FK-233-XA"

        # Clients réellement imprimés (on enlève le dépôt)
        codes_to_render = [c for c in (tour["nodes"] or []) if c and c.upper() != "FRESH DISTRIB"]

        # === Totaux calculés sur ces mêmes clients ===
        nb_resto = len(codes_to_render)
        p_total  = sum((metrics.get(c, {}).get("Poids",   0.0) or 0.0) for c in codes_to_render)
        c_total  = sum((metrics.get(c, {}).get("Cartons", 0.0) or 0.0) for c in codes_to_render)

        # --- En-tête jaune ---
        header_tbl = Table([[Paragraph(f"{tour_id}", style_title),
                             Paragraph(f"{veh} / {chauffeur}", style_hdr_right)]],
                           colWidths=[10.8*cm, 7.2*cm])
        header_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.yellow),
            ("BOX",        (0,0), (-1,0), 1, colors.black),
            ("VALIGN",     (0,0), (-1,0), "MIDDLE"),
            ("LEFTPADDING",(0,0), (-1,-1), 6),
            ("RIGHTPADDING",(0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0), (-1,-1), 6),
        ]))
        flow.append(header_tbl)
        flow.append(Spacer(1, 0.25*cm))

        # --- Synthèse (PDT / Cartons / Restaurants) ---
        synth = Table([[Paragraph(f"PDT : {int(round(p_total))} Kgs dont : 0 filets", style_line_big),
                        Paragraph(f"Boulangerie {int(round(c_total))} Cartons", style_line_big),
                        Paragraph(f"Restaurants : {nb_resto}", style_line_big)]],
                      colWidths=[8.0*cm, 5.0*cm, 5.0*cm])
        synth.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
            ("BOX",        (0,0), (-1,0), 1, colors.black),
            ("VALIGN",     (0,0), (-1,0), "MIDDLE"),
            ("LEFTPADDING",(0,0), (-1,-1), 6),
            ("RIGHTPADDING",(0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]))
        flow.append(synth)
        flow.append(Spacer(1, 0.15*cm))

        # --- Lignes clients (conforme au modèle + cadre + espace) ---
        for code in codes_to_render:
            g = geo_map.get(code, {})
            adr_full = (g.get("Adresse") or "").strip()

            # Nom du client : priorité à 'Nom', sinon tentative depuis le début de l'adresse
            name_val = g.get("Nom") or g.get("Nom client") or g.get("Raison sociale")
            adr_clean = adr_full
            if not name_val and adr_full:
                parts = [p.strip() for p in adr_full.split(",", 1)]
                if len(parts) == 2:
                    name_val, adr_clean = parts[0], parts[1]
                else:
                    name_val = None
            nom_affiche = name_val or code
            adr_affiche = adr_clean or ""

            m  = metrics.get(code, {})
            kg = int(round(m.get("Poids",   0.0) or 0.0))
            ct = int(round(m.get("Cartons", 0.0) or 0.0))

            # Largeurs colonnes
            col_w = [10.5*cm, 7.5*cm]

            # 1) LIGNE GRISE : nom seul
            row1 = Table([[Paragraph(nom_affiche, style_client),
                           Paragraph("", styles["Normal"])]],
                         colWidths=col_w)
            row1.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), colors.lightgrey),
                ("BOX",          (0,0), (-1,-1), 1, colors.black),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("RIGHTPADDING", (0,0), (-1,-1), 6),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ]))

            # 2) LIGNE BLANCHE : PDT/Cartons + adresse
            l2 = f"PDT : {kg} Kgs dont : 0 filets - Boulangerie {ct} Cartons"
            row2 = Table([[Paragraph(l2, style_line_small),
                           Paragraph(adr_affiche, styles["Normal"])]],
                         colWidths=col_w)
            row2.setStyle(TableStyle([
                ("BOX",          (0,0), (-1,-1), 1, colors.black),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("RIGHTPADDING", (0,0), (-1,-1), 6),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ]))

            # Espace entre clients
            flow.extend([row1, row2, Spacer(1, 0.15*cm)])

        # Trait final + saut de page
        flow.append(Table([[""]], colWidths=[18.0*cm],
                          style=TableStyle([("LINEBELOW", (0,0), (-1,0), 1, colors.black)])))
        if k < len(assigned):
            flow.append(PageBreak())

    doc.build(flow)
    buf.seek(0)
    return buf

# =========================================================
#                    CONFIG / STYLES
# =========================================================
st.set_page_config(layout="wide", page_title="Gestion & optimisation des tournées")
inject_brand_css()  # <= ICI : en dernier pour qu'il prenne le dessus

st.markdown(
    """
    <style>
      .css-18e3th9,
      .reportview-container .main > div:first-child { padding-top: 0 !important; margin-top: 0 !important; }
      .css-18e3th9 .main,
      .reportview-container .main { background: linear-gradient(to bottom, #e8eff5 0%, #ffffff 100%); }
      .css-1d391kg { background-color: #fafafa; padding-top: 0 !important; }
      .stSidebar .st-expander, .stExpander {
        background: white; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); padding: 0.5rem 1rem;
      }
      .sidebar-logo { display: flex; justify-content: center; margin-bottom: 1rem; }
    </style>
    """,
    unsafe_allow_html=True
)

# =========================================================
#                         HEADER
# =========================================================
# ⬇️ imports
from datetime import datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo   # Python ≥ 3.9
    PARIS_TZ = ZoneInfo("Europe/Paris")
except Exception:
    PARIS_TZ = None  # fallback si zoneinfo indispo (ex: manque du package tzdata)

def now_france_str(fmt: str = "%d/%m/%Y – %H:%M:%S") -> str:
    """Renvoie la date/heure en France (Europe/Paris), DST auto."""
    # 1) zoneinfo (recommandé)
    if PARIS_TZ is not None:
        return datetime.now(PARIS_TZ).strftime(fmt)
    # 2) fallback pytz si installé
    try:
        import pytz
        return datetime.now(pytz.timezone("Europe/Paris")).strftime(fmt)
    except Exception:
        # 3) fallback très simple (approx.) : UTC+2 (été) / adaptez si besoin
        return (datetime.utcnow().replace(tzinfo=timezone.utc) + timedelta(hours=2)).strftime(fmt)

inject_brand_css()

now = now_france_str()

# — Carte orange centrée
st.markdown(
    f"""
    <div class="welcome-wrap">
      <div class="welcome-card">
        <h2>👋 Bonjour et bienvenue !</h2>
        <p>📅 {now}</p>
      </div>
    </div>
    """,
    unsafe_allow_html=True
)

# — Titre principal centré
st.markdown('<h1 class="page-title">📋 Gestion clients & optimisation des tournées</h1>',
            unsafe_allow_html=True)
st.markdown("---")


st.markdown("""
**Mode d'emploi :**
- Les fichiers sont **chargés automatiquement depuis Google Drive** au lancement.
- **Optimiser** : onglet **🚚 Optimisation** → *Lancer l'optimisation* (**distance**).
- **Ajouter un client** : onglet **➕ Ajouter un client** → matrice & géocodage mis à jour **sur Drive**.
""")

# =========================================================
#             CHARGEMENT AUTOMATIQUE DEPUIS DRIVE
# =========================================================
# =========================================================
#             CHARGEMENT AUTOMATIQUE DEPUIS DRIVE
# =========================================================
REQUIRED_KEYS = ["distances", "geocodage", "vehicules", "chauffeurs"]

def _to_buf(b: bytes) -> BytesIO:
    buf = BytesIO(b); buf.seek(0); return buf

def load_from_drive_into_session():
    drive_cfg = st.secrets.get("drive", {})
    missing = [k for k in REQUIRED_KEYS if k not in drive_cfg]
    if missing:
        st.error(f"Secrets manquants dans [drive]: {', '.join(missing)}")
        st.stop()

    # Charge dist/geo/veh/chauff (obligatoires)
    filenames = {
        "dist":   drive_cfg["distances"],
        "geo":    drive_cfg["geocodage"],
        "veh":    drive_cfg["vehicules"],
        "chauff": drive_cfg["chauffeurs"],
    }
    for key, name in filenames.items():
        try:
            content = drive_download(name)
            st.session_state[f"{key}_buf"]  = _to_buf(content)
            st.session_state[f"{key}_name"] = name
        except Exception as e:
            st.session_state[f"{key}_buf"]  = None
            st.session_state[f"{key}_name"] = name
            st.sidebar.error(f"❌ Impossible de télécharger « {name} » depuis Drive : {e}")

    # Commandes = optionnel : on n’échoue pas si absent
    # Commandes = optionnel : ne rien afficher si absent
    orders_name = drive_cfg.get("commandes")
    st.session_state["orders_buf"]  = None
    st.session_state["orders_name"] = orders_name
    if orders_name:
        try:
            content = drive_download(orders_name)
            st.session_state["orders_buf"] = _to_buf(content)
        except Exception:
            # silencieux : pas de message dans la sidebar
            st.session_state["orders_buf"] = None
            pass


# 1ère initialisation (une seule fois)
if "initialized" not in st.session_state:
    for k in ["dist","geo","veh","chauff","orders"]:
        st.session_state.setdefault(f"{k}_buf",  None)
        st.session_state.setdefault(f"{k}_name", None)
    load_from_drive_into_session()
    st.session_state.initialized = True


# Accès rapide aux buffers (réutilisés dans les onglets)
dist_file   = st.session_state.get("dist_buf")
geo_file    = st.session_state.get("geo_buf")
orders_file = st.session_state.get("orders_buf")
veh_file    = st.session_state.get("veh_buf")
chauff_file = st.session_state.get("chauff_buf")
# ⚠️ Lis les Excel dans les onglets qui en ont besoin (ne pas lire ici).

# =========================================================
#                 SIDEBAR : DONNÉES / RECHARGER
# =========================================================
APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "assets" / "company_logo.png"

if LOGO_PATH.exists():
    st.sidebar.image(str(LOGO_PATH), use_container_width=True)
else:
    st.sidebar.warning("Logo introuvable (assets/company_logo.png)")
st.sidebar.header("📂 Données")

def _all_loaded():
    return all(st.session_state.get(k) for k in ["dist_buf","geo_buf","veh_buf","chauff_buf"])


if _all_loaded():
    st.sidebar.success("✅ Données chargées. Pour recharger, cliquez sur le bouton.")
else:
    st.sidebar.warning("⚠️ Données incomplètes. Cliquez sur le bouton pour recharger.")

if st.sidebar.button("🔄 Recharger depuis Drive"):
    load_from_drive_into_session()
    st.sidebar.success("✅ Données rechargées.")
    st.rerun()



# ====================== ONGLETS PRINCIPAUX ======================
# ====================== ONGLETS PRINCIPAUX ======================
# AVANT : tab_opt, tab_add = st.tabs(["🚚 Optimisation", "➕ Ajouter un client"])
tab_opt, tab_add, tab_drivers = st.tabs([
    "🚚 Optimisation",
    "➕ Ajouter un client",
    "👷 Gestion des chauffeurs"
])

# =========================================================
#                        ONGLET OPTIMISATION
# =========================================================
# =========================================================
#                        ONGLET OPTIMISATION
# =========================================================
with tab_opt:
    st.subheader("Paramètres d'optimisation")
    # --------- Source commandes : import local prioritaire ----------

    with st.expander("📥 Importer le fichier de commandes du jour (xlsx/csv)", expanded=True):
        up = st.file_uploader(
            "Glissez-déposez ou cliquez pour sélectionner votre fichier (.xlsx, .xls, .csv)",
            type=["xlsx", "xls", "csv"],
            key="orders_upload",
            help="Colonnes attendues : Code client, Libellé, Quantité, Unité, Adresse, Code postal, Ville",
        )
    
        # (Optionnel) Indiquer la source actuelle ; supprime ces 2 lignes si tu ne veux rien afficher
        src = "📎 Import local" if st.session_state.get("orders_source") == "upload" \
              else f"Drive : {st.session_state.get('orders_name') or st.secrets['drive'].get('commandes','(non défini)')}"
        st.caption(f"**Source actuelle des commandes** : {src}")
    
        if up is not None:
            try:
                buf, df_preview, missing = _uploaded_orders_to_excelbuf(up)
                if missing:
                    st.error("Colonnes manquantes : " + ", ".join(missing))
                    st.markdown("Aperçu pour diagnostic :")
                    st.dataframe(df_preview.head(10), use_container_width=True)
                else:
                    st.session_state["orders_buf"] = buf
                    st.session_state["orders_name"] = up.name
                    st.session_state["orders_source"] = "upload"
                    st.success(f"✅ « {up.name} » importé. Il sera utilisé pour l'optimisation.")
                    buf.seek(0)
                    st.dataframe(pd.read_excel(buf).head(10), use_container_width=True)
                    buf.seek(0)
            except Exception as e:
                st.error(f"Échec import : {e}")


    # Indisponibilités
    unv_veh, unv_ch = [], []

    # ---------- Véhicules indisponibles ----------
    if veh_file:
        veh_file.seek(0)
        try:
            dfv = pd.read_excel(veh_file, skiprows=1)
            unv_veh = unavail_multiselect(
                "🚫 Véhicules indisponibles",
                dfv["Véhicule"].dropna().astype(str).unique().tolist(),
                key="veh_unavail"
            )
        finally:
            veh_file.seek(0)

    # ---------- Chauffeurs indisponibles + remplaçants (temporaires même véhicule) ----------
    # Chauffeurs indisponibles
    # ---------- Chauffeurs indisponibles + remplaçants (temporaires même véhicule) ----------
    if chauff_file:
        chauff_file.seek(0)
        try:
            dfc = pd.read_excel(chauff_file, sheet_name="Liste")
    
            # Nom complet
            dfc["Nom Complet"] = (
                dfc["Nom"].astype(str).fillna("") + " " + dfc["Prénom"].astype(str).fillna("")
            ).str.strip()
    
            # Helper: statut temporaire ?
            def _is_temp_statut(x) -> bool:
                s = unicodedata.normalize("NFKD", str(x or "")).encode("ascii", "ignore").decode("ascii").lower()
                # couvre: "temp", "temporaire", "intérim", "interim", "extra"
                return any(k in s for k in ("temp", "interim", "interimaire", "extra"))
    
            # 1) Liste pour "Chauffeurs indisponibles" = PERMANENTS uniquement
            if "Statut" in dfc.columns:
                df_perm = dfc.loc[~dfc["Statut"].apply(_is_temp_statut)].copy()
            else:
                df_perm = dfc.copy()  # si colonne manquante, on garde tout
    
            all_ch_perm = sorted([n for n in df_perm["Nom Complet"].dropna().unique().tolist() if n])
            unv_ch = unavail_multiselect(
                "🚫 Chauffeurs indisponibles (permanents uniquement)",
                all_ch_perm,
                key="ch_unavail"
            )
    
            # 2) Remplaçants = TEMPORAIRES sur le même véhicule
            selected_replacements = {}
            restrict_to_selected = False
    
            if "Statut" in dfc.columns and unv_ch:
                df_temp = dfc.loc[dfc["Statut"].apply(_is_temp_statut)].copy()
                if not df_temp.empty:
                    st.markdown("#### 🤝 Remplaçants (temporaires **même véhicule**)")
    
                    veh_by_name = dict(zip(dfc["Nom Complet"], dfc["Véhicule affecté"]))
                    already_taken = set()
    
                    for i, ch in enumerate(unv_ch):
                        veh = veh_by_name.get(ch, "")
    
                        if veh in (unv_veh or []):
                            alert_white_red(f"• <b>{ch}</b> → véhicule <b>{veh}</b> indisponible : pas de remplaçant proposé.")
                            continue
    
                        same_veh_temps = df_temp.loc[df_temp["Véhicule affecté"].astype(str) == str(veh), "Nom Complet"].tolist()
                        same_veh_temps = [t for t in same_veh_temps if t not in already_taken]
    
                        if not same_veh_temps:
                            alert_white_red(f"• <b>{ch}</b> → aucun <b>temporaire</b> disponible sur le véhicule <b>{veh}</b>.")
                            continue
    
                        options = ["— Aucun —"] + same_veh_temps
                        rep = st.selectbox(
                            f"Remplaçant pour **{ch}** (véhicule {veh})",
                            options,
                            index=1 if len(options) > 1 else 0,
                            key=f"rep_sameveh_{i}"
                        )
                        if rep != "— Aucun —":
                            selected_replacements[ch] = rep
                            already_taken.add(rep)
                else:
                    alert_white_red("Aucun chauffeur temporaire dans la feuille <b>Liste</b>.")
        finally:
            chauff_file.seek(0)


        # ------------------- Lancer l’optimisation -------------------
        if st.button("🚀 Lancer l'optimisation"):
            if not all([veh_file, chauff_file, dist_file, geo_file]):
                st.error("⚠️ Fichiers manquants sur Drive (voir panneau de gauche).")
            elif not orders_file:
                st.error("⚠️ Fichier de commandes manquant.")
            else:
                # Construire la liste finale des chauffeurs à exclure
                unv_ch_final = list(unv_ch)
                
                # AUTO : si au moins un remplaçant est choisi, exclure tous les autres temporaires
                if dfc is not None and "Statut" in dfc.columns and len(selected_replacements) > 0:
                    temp_all = set(
                        dfc.loc[dfc["Statut"].astype(str).str.lower().str.contains("temp"),
                                "Nom Complet"].tolist()
                    )
                    keep = set(selected_replacements.values())
                    extra_unavailable = list(temp_all - keep)
                    unv_ch_final.extend(extra_unavailable)


                # Petit résumé des remplacements
                if selected_replacements:
                    pairs = "\n".join([f"- {u} → {r}" for u, r in selected_replacements.items()])
                    st.info(f"Remplaçants choisis :\n{pairs}")

                with st.spinner("Optimisation en cours…"):
                    st.session_state.dist_buf.seek(0)
                    orders_file.seek(0); veh_file.seek(0); chauff_file.seek(0)

                    result, out_xl = run_optimization(
                        st.session_state.dist_buf,
                        orders_file, veh_file, chauff_file,
                        "distance",
                        unavailable_vehicles=unv_veh,
                        unavailable_chauffeurs=unv_ch_final,
                    )

                    if out_xl is None:
                        st.error("⚠️ Aucune solution trouvée.")
                    else:
                        st.session_state.result      = result
                        st.session_state.output_xlsx = out_xl

        # ===== Résultats + MàJ Chauffeurs + PDF + Carte =====
        if st.session_state.get("result"):
            st.success("✅ Optimisation terminée")

            st.text_area("Résumé des tournées", st.session_state.result, height=300)

            # MàJ du classeur Chauffeurs (Kilométrage + cumul) sur Drive
            # MàJ du classeur Chauffeurs (Kilométrage + cumul) sur Drive
            if chauff_file:
                try:
                    # 1) lire la feuille 'Kilométrage' du fichier résultat
                    st.session_state.output_xlsx.seek(0)
                    df_new = pd.read_excel(
                        st.session_state.output_xlsx,
                        sheet_name="Kilométrage",
                        engine="openpyxl",
                        dtype={0: str, 1: str, 2: str, 3: float}  # date, chauffeur, tournée, distance
                    )

                    if df_new.empty:
                        st.info("Aucune nouvelle ligne Kilométrage trouvée dans le résultat.")
                    else:
                        # normaliser la date au format YYYY-MM-DD (au cas où)
                        def _norm_date(s: str) -> str:
                            s = str(s)
                            if " " in s:
                                s = s.split(" ", 1)[0]
                            try:
                                return pd.to_datetime(s, dayfirst=True).date().isoformat()
                            except Exception:
                                return s
                        df_new.iloc[:, 0] = df_new.iloc[:, 0].map(_norm_date)

                        # filtrer d'éventuelles lignes d'entêtes répétées
                        headers = [h.lower().strip() for h in df_new.columns]
                        def _is_header_like(row):
                            return all(str(row[i]).lower().strip() == headers[i] for i in range(len(headers)))
                        df_new = df_new.loc[~df_new.apply(_is_header_like, axis=1)]

                        # tuples (date, chauffeur, tournée, distance)
                        new_entries = [tuple(x) for x in df_new.iloc[:, :4].itertuples(index=False, name=None)]

                        if not new_entries:
                            st.info("Rien à intégrer dans 'Kilométrage'.")
                        else:
                            # 2) charger le classeur Chauffeurs, mettre à jour, ré-uploader
                            chauff_file.seek(0)
                            wb_ch = load_workbook(chauff_file)

                            update_kilometrage_file(wb_ch, new_entries)  # ⬅️ remplace les dates du jour et ajoute
                            update_distance_parcourue(wb_ch)             # ⬅️ recalcule les cumuls

                            buf_ch = BytesIO()
                            wb_ch.save(buf_ch); buf_ch.seek(0)

                            drive_upload(st.secrets["drive"]["chauffeurs"], buf_ch.getvalue())
                            st.session_state.chauff_buf = BytesIO(buf_ch.getvalue()); st.session_state.chauff_buf.seek(0)

                            st.success(f"✅ Le fichier des Chauffeurs a été mis à jour.")
                except Exception as e:
                    st.error(f"❌ Échec mise à jour du classeur Chauffeurs sur Drive : {e}")


            # ---------- PDF multi-tournées ----------
            if _PDF_IMPORT_ERROR:
                st.error(f"Module ReportLab requis pour le PDF : {_PDF_IMPORT_ERROR}\nInstalle : pip install reportlab")
            else:
                try:
                    assigned = []
                    for i, blk in enumerate(st.session_state.result.split("Tournée ")[1:]):
                        lines = blk.split("\n")
                        try:
                            tid = f"Tournée {i+1}"
                            ch, vh = _extract_chauffeur_vehicle(lines[0])
                            nodes = next(l for l in lines if "Clients" in l).split(":")[1].split("->")
                            nodes = [n.strip() for n in nodes]
                            assigned.append({"tour_id": tid, "chauffeur": ch, "vehicle": vh.strip(), "nodes": nodes})
                        except StopIteration:
                            pass

                    st.session_state.geo_buf.seek(0)
                    df_geo_for_pdf = pd.read_excel(st.session_state.geo_buf)
                    orders_file.seek(0)
                    df_orders_for_pdf = pd.read_excel(orders_file)

                    pdf_bytes = _generate_pdf_tours(assigned, df_geo_for_pdf, df_orders_for_pdf, st.session_state.result)
                    st.download_button(
                        "📄 Télécharger PDF des tournées",
                        data=pdf_bytes,
                        file_name = f"Tournées_{datetime.now(PARIS_TZ):%Y-%m-%d}.pdf" if PARIS_TZ else f"Tournées_{datetime.utcnow():%Y-%m-%d}.pdf",
                        mime="application/pdf",
                        key="dl_pdf"
                    )

                
                except Exception as e:
                    st.error(f"Erreur génération PDF : {e}")

            # --- Carte interactive ---
            # --- Carte interactive (filtre = Chauffeur uniquement) ---
            if st.session_state.get("geo_buf") and st.session_state.get("result"):
                st.session_state.geo_buf.seek(0)
                df_coords = pd.read_excel(st.session_state.geo_buf)

                # colonne code client / dépôt
                code_col = next((c for c in df_coords.columns if "Code" in c and "tiers" in c), None)
                df_coords["_code"] = df_coords.get(code_col, "FRESH DISTRIB").fillna("FRESH DISTRIB")

                # nom du client (prend "Nom du client" si dispo, sinon "Nom")
                name_col = "Nom du client" if "Nom du client" in df_coords.columns else ("Nom" if "Nom" in df_coords.columns else None)
                cols = ["Latitude", "Longitude"] + ([name_col] if name_col else [])
                coord_dict = df_coords.set_index("_code")[cols].to_dict("index")

                # Reparsage des tournées depuis le résumé texte (inchangé)
                assigned_map = []
                for i, blk in enumerate(st.session_state.result.split("Tournée ")[1:]):
                    lines = blk.split("\n")
                    try:
                        tour_id = f"Tournée {i+1}"
                        # ligne 1: "Tournée i : Chauffeur XYZ via VEHICULE"
                        header = lines[0]
                        # Chauffeur entre " : " et " via "
                        ch = header.split(" : ", 1)[1].split(" via ")[0].strip()
                        vh = header.split(" via ", 1)[1].strip()
                        # ligne "Clients  : A -> B -> ... "
                        nodes = next(l for l in lines if "Clients" in l).split(":")[1].split("->")
                        nodes = [n.strip() for n in nodes]
                        assigned_map.append({"tour_id": tour_id, "chauffeur": ch, "vehicle": vh, "nodes": nodes})
                    except Exception:
                        pass

                # ---- UI : un seul filtre Chauffeur ----
                col_map, col_fil = st.columns([3, 1])
                with col_fil:
                    st.markdown("### 🔎 Filtres d'affichage")
                    chauffeurs = ["Tous"] + sorted({t["chauffeur"] for t in assigned_map})
                    sel_ch = st.selectbox("👷 Chauffeur", chauffeurs, index=0)

                    # Affiche le véhicule quand un chauffeur est sélectionné
                    if sel_ch != "Tous":
                        vehs = sorted({t["vehicle"] for t in assigned_map if t["chauffeur"] == sel_ch})
                        if vehs:
                            st.markdown(
                                f"""
                                <div class="force-black"
                                     style="margin-top:6px;padding:8px 10px;border-left:4px solid #0ea5e9;background:#f0f9ff;">
                                  <span style="font-weight:700;">Véhicule utilisé :</span> {', '.join(vehs)}
                                </div>
                                """,
                                unsafe_allow_html=True
                            )
                    


                    filtered = [t for t in assigned_map if sel_ch == "Tous" or t["chauffeur"] == sel_ch]

                # ---- Carte ----
                with col_map:
                    lats, lons = [], []
                    for tour in filtered:
                        for code in tour["nodes"]:
                            if code in coord_dict:
                                lats.append(coord_dict[code]['Latitude'])
                                lons.append(coord_dict[code]['Longitude'])
                    if lats:
                        center = [sum(lats)/len(lats), sum(lons)/len(lons)]
                        m = folium.Map(location=center, zoom_start=11)
                        colors = ["red","blue","green","purple","orange","darkred","cadetblue","darkgreen","darkpurple","lightred"]
                        for i, tour in enumerate(filtered):
                            color = colors[i % len(colors)]
                            path = []
                            for code in tour["nodes"]:
                                if code not in coord_dict:
                                    continue
                                pt = (coord_dict[code]['Latitude'], coord_dict[code]['Longitude'])
                                path.append(pt)
                                label = (f"{code} : {coord_dict[code].get(name_col)}"
                                        if name_col and code != "FRESH DISTRIB"
                                        else (coord_dict[code].get(name_col) or "FRESH DISTRIB"))
                                if code == "FRESH DISTRIB":
                                    folium.Marker(pt, icon=folium.Icon(color="darkblue", icon="home", prefix="fa"),
                                                popup=label).add_to(m)
                                else:
                                    folium.CircleMarker(pt, radius=6, color=color, fill=True,
                                                        fill_color="white", fill_opacity=1, popup=label).add_to(m)
                            if path:
                                folium.PolyLine(path, color=color, weight=3, opacity=0.8).add_to(m)
                        st.subheader("📜 Carte interactive des tournées")
                        folium_static(m)
                    else:
                        st.info("Aucun point à afficher pour ce filtre.")


# =========================================================
#                ONGLET GESTION DES CHAUFFEURS
# =========================================================
# =========================================================
#                ONGLET GESTION DES CHAUFFEURS
# =========================================================
with tab_drivers:
    st.header("🚚 Gestion des chauffeurs")

    # --- helpers locaux (reset différé + utilitaires xlsx) ---
    def _deferred_reset(flag_key: str, keys: list):
        """Si flag actif, supprimer les clés AVANT création des widgets, puis éteindre le flag."""
        if st.session_state.get(flag_key):
            for k in keys:
                st.session_state.pop(k, None)
            st.session_state[flag_key] = False

    def _get_ws(_wb, wanted="Liste"):
        import re as _re
        names = {_re.sub(r"\s+", "", s.lower()): s for s in _wb.sheetnames}
        key = _re.sub(r"\s+", "", wanted.lower())
        return _wb[names[key]] if key in names else _wb.create_sheet(wanted)

    def _first_empty_row_compact(_ws, cols, start=2):
        """Première vraie ligne vide en vérifiant seulement les colonnes clés."""
        last_real = start - 1
        maxr = _ws.max_row or (start - 1)
        for r in range(start, maxr + 1):
            if any(str(_ws.cell(row=r, column=c).value or "").strip() for c in cols):
                last_real = r
        return last_real + 1

    sub_tab_add, sub_tab_del = st.tabs([
        "➕ Ajouter un chauffeur",
        "🗑️ Supprimer définitivement un chauffeur"
    ])

    # -------------------- ➕ AJOUTER UN CHAUFFEUR --------------------
    with sub_tab_add:
        # reset différé des champs d'ajout
        _deferred_reset("_reset_add_form", ["nom_input", "prenom_input", "vehicule_select", "confirm_add", "statut_select"])

        col_a, col_b = st.columns(2)
        with col_a:
            nom = st.text_input("Nom", "", key="nom_input")
        with col_b:
            prenom = st.text_input("Prénom", "", key="prenom_input")

        # Liste des véhicules depuis vehicles.xlsx (SEULE source)
        vehicule_options = ["— Sélectionner —"]
        if veh_file:
            try:
                veh_file.seek(0)
                dfv = pd.read_excel(veh_file, skiprows=1)
                if "Véhicule" in dfv.columns:
                    vehicule_options += [
                        v for v in dfv["Véhicule"].dropna().astype(str).unique().tolist() if v.strip()
                    ]
            except Exception as e:
                st.warning(f"Impossible de lire le fichier Véhicules : {e}")

        chosen_vehicle = st.selectbox(
            "Véhicule affecté (depuis la liste)",
            vehicule_options,
            index=0,
            key="vehicule_select"
        )

        # NOUVEAU : Statut Permanent / Temporaire
        statut = st.selectbox("Statut du chauffeur", ["Permanent", "Temporaire"], index=0, key="statut_select")

        # -------- FORMULAIRE AVEC CONFIRMATION + PROGRESSION --------
        with st.form("form_add_driver", clear_on_submit=False):
            confirm = st.selectbox("Confirmer l'ajout de ce chauffeur ?", ["Non", "Oui"], index=0, key="confirm_add")
            submitted = st.form_submit_button("💾 Cliquez pour enregistrer le chauffeur")

            if submitted:
                # Validations
                if not nom.strip() or not prenom.strip():
                    st.error("Complète **Nom** et **Prénom**."); st.stop()
                if chosen_vehicle == "— Sélectionner —":
                    st.error("Sélectionne un **Véhicule** dans la liste."); st.stop()
                if st.session_state.get("confirm_add") != "Oui":
                    st.error("Merci de confirmer l'ajout (choisir **Oui**)."); st.stop()
                if not st.session_state.get("chauff_buf"):
                    st.error("Fichier Chauffeurs introuvable."); st.stop()

                try:
                    prog = st.progress(0, text="Initialisation…")

                    # 1) Ouvrir le classeur existant
                    prog.progress(15, text="Ouverture du classeur…")
                    st.session_state["chauff_buf"].seek(0)
                    original_bytes = st.session_state["chauff_buf"].read()
                    from openpyxl import load_workbook
                    wb = load_workbook(BytesIO(original_bytes))

                    # 2) Feuille 'Liste'
                    ws = _get_ws(wb, "Liste")

                    # 3) Garantir les colonnes nécessaires, sans écraser les autres
                    prog.progress(35, text="Préparation de la feuille…")
                    # Nettoyage d'une ancienne colonne 'Actif' si présente
                    headers_now = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                    if "Actif" in headers_now:
                        col_act = headers_now.index("Actif") + 1
                        ws.delete_cols(col_act, 1)

                    # Helper: créer la colonne si elle n'existe pas, sinon renvoyer son index
                    def _ensure_col(name):
                        hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                        if name in hdrs:
                            return hdrs.index(name) + 1
                        col = ws.max_column + 1
                        ws.cell(1, col, name)
                        return col

                    col_nom   = _ensure_col("Nom")
                    col_pre   = _ensure_col("Prénom")
                    col_veh   = _ensure_col("Véhicule affecté")
                    col_stat  = _ensure_col("Statut")     # <-- NOUVELLE COLONNE
                    # (Les autres colonnes comme "Distance parcourue (km)" restent en place)

                    # 4) Écriture compacte
                    prog.progress(55, text="Écriture de la nouvelle ligne…")
                    target_row = _first_empty_row_compact(ws, [col_nom, col_pre, col_veh], start=2)
                    ws.cell(target_row, col_nom, nom.strip())
                    ws.cell(target_row, col_pre, prenom.strip())
                    ws.cell(target_row, col_veh, chosen_vehicle)
                    ws.cell(target_row, col_stat, statut)  # <-- écrit le statut

                    # 5) Sauvegarde + Upload
                    prog.progress(75, text="Sauvegarde du fichier…")
                    out = BytesIO(); wb.save(out); out.seek(0)

                    prog.progress(85, text="Envoi vers Google Drive…")
                    drive_upload(st.secrets["drive"]["chauffeurs"], out.getvalue())

                    # 6) Re-télécharger & rebind
                    prog.progress(95, text="Vérification de l'écriture…")
                    b = drive_download(st.secrets["drive"]["chauffeurs"])
                    tmp = BytesIO(b); tmp.seek(0)
                    df_check = pd.read_excel(tmp, sheet_name="Liste")

                    st.session_state["chauff_buf"] = BytesIO(b)
                    st.session_state["chauff_buf"].seek(0)

                    prog.progress(100, text="Terminé ✅")
                    st.success(f"✅ {nom} {prenom} ajouté à la liste (ligne {target_row}).")

                    # Affichage de la liste mise à jour (sera aussi affichée après rerun)
                    st.markdown("### 📃 Liste des chauffeurs (mise à jour)")
                    st.dataframe(df_check, use_container_width=True)

                    # Reset différé des champs + rerun
                    st.session_state["_reset_add_form"] = True
                    st.rerun()

                except Exception as e:
                    st.error(f"Erreur lors de l'enregistrement : {e}")

        # Afficher la liste courante en dessous (utile après rerun)
        if st.session_state.get("chauff_buf"):
            try:
                st.session_state["chauff_buf"].seek(0)
                _df_list = pd.read_excel(st.session_state["chauff_buf"], sheet_name="Liste")
                st.markdown("### 📃 Liste des chauffeurs (courante)")
                st.dataframe(_df_list, use_container_width=True)
            except Exception as _e:
                st.warning(f"Impossible d'afficher la liste des chauffeurs : {_e}")

    # -------------------- 🗑️ SUPPRIMER DÉFINITIVEMENT --------------------
    with sub_tab_del:
        _deferred_reset("_reset_del_form", ["del_choice", "del_ack", "del_text"])

        _chauff_buf = st.session_state.get("chauff_buf")
        if not _chauff_buf:
            st.info("Fichier Chauffeurs introuvable.")
        else:
            _chauff_buf.seek(0)
            try:
                df_ch = pd.read_excel(_chauff_buf, sheet_name="Liste")
            except Exception as e:
                st.error(f"Impossible de lire la feuille 'Liste' : {e}")
                df_ch = pd.DataFrame(columns=["Nom","Prénom","Véhicule affecté","Statut"])

            if df_ch.empty or ("Nom" not in df_ch.columns) or ("Prénom" not in df_ch.columns):
                st.info("Aucun chauffeur exploitable dans la feuille 'Liste'.")
            else:
                import re
                def _norm(x: str) -> str:
                    s = str(x or "")
                    s = re.sub(r"\s+", " ", s).strip().rstrip(":").lower()
                    return s

                # Options uniques Nom+Prénom
                seen = set(); opts = []
                for _, r in df_ch.iterrows():
                    key = (_norm(r.get("Nom")), _norm(r.get("Prénom")))
                    if key not in seen and any(key):
                        seen.add(key)
                        label = f"{str(r.get('Nom') or '').strip()} {str(r.get('Prénom') or '').strip()}".strip()
                        if label:
                            opts.append({"label": label, "nom": r.get("Nom"), "pre": r.get("Prénom")})

                choice = st.selectbox(
                    "Sélectionner le chauffeur à supprimer",
                    ["— Aucun —"] + [o["label"] for o in opts],
                    index=0,
                    key="del_choice"
                )

                if st.session_state.get("del_choice") != "— Aucun —":
                    sel = next((o for o in opts if o["label"] == st.session_state["del_choice"]), None)
                    sel_nom, sel_pre = sel["nom"], sel["pre"]

                    mask = df_ch.apply(lambda r: _norm(r.get("Nom")) == _norm(sel_nom)
                                                and _norm(r.get("Prénom")) == _norm(sel_pre), axis=1)
                    st.markdown("**Entrées correspondantes dans 'Liste' :**")
                    show_cols = [c for c in ["Nom","Prénom","Véhicule affecté","Statut"] if c in df_ch.columns]
                    st.dataframe(df_ch.loc[mask, show_cols], use_container_width=True)

                    st.markdown(
                        "<span style='color:red; font-weight:600;'>"
                        "⚠️ Action irréversible : seules les lignes de la feuille 'Liste' seront supprimées. "
                        "Les autres feuilles (dont 'Kilométrage') ne seront pas modifiées."
                        "</span>", unsafe_allow_html=True
                    )
                    ok = st.checkbox("Je comprends que cette action est irréversible.", key="del_ack")
                    txt = st.text_input("Tapez SUPPRIMER pour confirmer", "", key="del_text")

                    if st.button("🗑️ Confirmer suppression"):
                        if not ok or st.session_state.get("del_text","").strip().upper() != "SUPPRIMER":
                            st.error("Confirmez en cochant la case et en tapant exactement SUPPRIMER.")
                        else:
                            try:
                                _chauff_buf.seek(0)
                                original = _chauff_buf.read()
                                from openpyxl import load_workbook
                                wb = load_workbook(BytesIO(original))
                                ws = _get_ws(wb, "Liste")

                                headers = [c.value for c in ws[1]]
                                col_nom = headers.index("Nom") + 1
                                col_pre = headers.index("Prénom") + 1

                                rows_to_delete = []
                                for r in range(2, ws.max_row + 1):
                                    if (_norm(ws.cell(r, col_nom).value) == _norm(sel_nom) and
                                        _norm(ws.cell(r, col_pre).value) == _norm(sel_pre)):
                                        rows_to_delete.append(r)

                                if not rows_to_delete:
                                    st.warning("Aucune ligne correspondante trouvée dans 'Liste'.")
                                else:
                                    for r in reversed(rows_to_delete):
                                        ws.delete_rows(r, 1)

                                    out = BytesIO(); wb.save(out); out.seek(0)
                                    drive_upload(st.secrets["drive"]["chauffeurs"], out.getvalue())

                                    st.session_state["chauff_buf"] = BytesIO(out.getvalue())
                                    st.session_state["chauff_buf"].seek(0)

                                    st.success(f"✅ {st.session_state['del_choice']} supprimé de 'Liste' ({len(rows_to_delete)} ligne(s)).")

                                    # Reset différé + rerun
                                    st.session_state["_reset_del_form"] = True
                                    st.rerun()

                            except Exception as e:
                                st.error(f"Erreur pendant la suppression : {e}")



# ==============================
# Fonction pour géocoder via Google
# ==============================
def geocode_google(address, api_key):
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {"address": address, "key": api_key}
    r = requests.get(url, params=params)
    r.raise_for_status()
    data = r.json()
    if data["status"] != "OK":
        raise ValueError(f"Google Geocoding API error: {data['status']}")
    loc = data["results"][0]["geometry"]["location"]
    return loc["lat"], loc["lng"]
# =========================================================
#                      ONGLET AJOUT CLIENT
# =========================================================
with tab_add:
    # Crée les colonnes d'abord : le titre sera dans la colonne de gauche
    col_left, col_right = st.columns([2.2, 1])

    with col_left:
        st.subheader("Ajouter un nouveau client")
        st.caption("Cette action met à jour **Géocodage** et **Matrice des distances**.")

        # Formulaire
        addr = st.text_input("Adresse complète", placeholder="Nom, adresse")
        code = st.text_input("Code (tiers)")
        add_btn = st.button("➕ Ajouter & mettre à jour Drive")

    # La carte s'affichera tout en haut de la colonne de droite (alignée avec le titre)
    with col_right:
        map_placeholder = st.empty()   # on remplit après géocodage

    if add_btn:
        if not (st.session_state.get("geo_buf") and st.session_state.get("dist_buf")):
            st.error("⚠️ Les fichiers Géocodage & Matrice doivent exister sur Drive (voir secrets).")
        elif not (addr and code):
            st.error("⚠️ Complétez adresse et code tiers.")
        else:
            # --- Géocodage avec Google Maps API ---
            api_key = st.secrets["google"]["api_key"]  # stockée dans secrets.toml
            url = "https://maps.googleapis.com/maps/api/geocode/json"
            params = {"address": addr, "key": api_key}
            r = requests.get(url, params=params)
            r.raise_for_status()
            data = r.json()

            if data.get("status") != "OK":
                st.error(f"Erreur de géocodage Google Maps : {data.get('status')}")
                st.stop()

            location = data["results"][0]["geometry"]["location"]
            lat = float(location["lat"])
            lon = float(location["lng"])

            print(f"[DEBUG] Géocodage OK - {addr} => lat={lat}, lon={lon}")

            # --- Afficher la mini-carte à droite (alignée avec le titre) ---
            try:
                m = folium.Map(location=[lat, lon], zoom_start=16)
                folium.Marker([lat, lon], popup=f"{code}<br>{addr}", tooltip="Nouveau client").add_to(m)
                with col_right:
                    folium_static(m, width=380, height=280)  # ajuste si besoin
            except Exception as _e:
                with col_right:
                    st.info(f"Prévisualisation carte non disponible : {_e}")

            # --- Chargement des fichiers ---
            st.session_state.geo_buf.seek(0)
            df_geo = pd.read_excel(st.session_state.geo_buf)
            df_geo["Code (tiers)"] = df_geo.get("Code (tiers)", pd.Series()).fillna("FRESH DISTRIB")

            st.session_state.dist_buf.seek(0)
            df_mat = pd.read_excel(st.session_state.dist_buf, index_col=0)

            # --- Ajout du nouveau client ---
            new_geo = {"Adresse": addr, "Code (tiers)": code, "Latitude": lat, "Longitude": lon}
            df_geo = pd.concat([df_geo, pd.DataFrame([new_geo])], ignore_index=True)
            df_geo["Code (tiers)"].fillna("FRESH DISTRIB", inplace=True)

            # --- Calcul des distances OSRM ---
            coords = {r["Code (tiers)"]:(r["Longitude"], r["Latitude"]) for _, r in df_geo.iterrows()}

            def route(o, d):
                url = f"http://router.project-osrm.org/route/v1/driving/{o[0]},{o[1]};{d[0]},{d[1]}"
                r = requests.get(url); r.raise_for_status()
                rt = r.json().get("routes")
                return ((rt[0]["distance"]/1000, rt[0]["duration"]/60) if rt else (0, 0))

            fw, bw, errors = {}, {}, 0
            total = len(coords)

            # Place les logs/progression sous le formulaire (colonne gauche)
            with col_left:
                prog = st.progress(0, text="Initialisation…")
                # Message rouge d'avertissement
                st.markdown(
                    "<span style='color:red; font-weight:bold;'>⚠️ Calcul des distances en cours… veuillez patienter et ne pas fermer ou recharger la page.</span>",
                    unsafe_allow_html=True
                )
                log  = st.empty()


            for idx, (c, (clon, clat)) in enumerate(coords.items(), start=1):
                try:
                    if c == code:
                        fw[c] = bw[c] = (0, 0)
                    else:
                        d1, t1 = route((lon, lat), (clon, clat)); time.sleep(0.2)
                        d2, t2 = route((clon, clat), (lon, lat)); time.sleep(0.2)
                        fw[c], bw[c] = (d1, t1), (d2, t2)

                    pct = int(idx / total * 100)
                    with col_left:
                        prog.progress(pct, text=f"Calcul {idx}/{total} : {code} ↔ {c} …")
                        log.markdown(
                            f"- `{code}→{c}` : **{fw[c][0]:.2f} km** / {fw[c][1]:.0f} min  |  "
                            f"`{c}→{code}` : **{bw[c][0]:.2f} km** / {bw[c][1]:.0f} min"
                        )
                except Exception as e:
                    errors += 1
                    fw[c] = fw.get(c, (0, 0))
                    bw[c] = bw.get(c, (0, 0))
                    with col_left:
                        prog.progress(int(idx / total * 100), text=f"⚠️ Erreur sur {c} (continuer)…")
                        log.warning(f"Erreur {code}↔{c} : {e}")

            with col_left:
                prog.empty()
                if errors:
                    st.warning(f"Terminé avec {errors} échec(s).")
                else:
                    st.success("Calcul des distances & durées terminé.")

            # --- Mise à jour de la matrice ---
            df_mat[code] = [f"{fw[k][0]:.2f} km / {fw[k][1]:.2f} min" for k in df_mat.index]
            new_row = {
                col: ("0 km / 0 min" if col == code else f"{bw[col][0]:.2f} km / {bw[col][1]:.2f} min")
                for col in df_mat.columns
            }
            df_mat.loc[code] = pd.Series(new_row)

            # --- Sauvegarde sur Drive ---
            geo_out = BytesIO(); df_geo.to_excel(geo_out, index=False); geo_out.seek(0)
            mat_out = BytesIO(); df_mat.to_excel(mat_out, index=True); mat_out.seek(0)
            try:
                drive_upload(st.secrets["drive"]["geocodage"], geo_out.getvalue())
                drive_upload(st.secrets["drive"]["distances"], mat_out.getvalue())
                with col_left:
                    st.success("✅ Géocodage & matrice mis à jour sur Drive.")
            except Exception as e:
                with col_left:
                    st.error(f"❌ Échec d'écriture sur Drive : {e}")






















