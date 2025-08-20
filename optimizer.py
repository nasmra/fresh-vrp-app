from collections import defaultdict, Counter
import os
import re
import math
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ortools.constraint_solver import pywrapcp, routing_enums_pb2

# --- Timezone France (optionnel)
try:
    from zoneinfo import ZoneInfo
    PARIS_TZ = ZoneInfo("Europe/Paris")
except Exception:
    PARIS_TZ = None


# ======================================================
# Utils
# ======================================================
def _norm(s): 
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()

def _norm_space(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()

def sanitize_sheet_name(name): 
    return re.sub(r'[\\/*?:\[\]]', '', name)[:31]


# ---------- Parsing robuste km / min ----------
def _parse_km_cell(x):
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)
    if isinstance(x, str):
        s = x.strip().lower().replace(",", ".")
        m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*km", s)
        if m:
            return float(m.group(1))
        try:
            return float(re.sub(r"[^\d\.]", "", s))
        except Exception:
            return 0.0
    return 0.0

def _parse_min_cell(x):
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)
    if isinstance(x, str):
        s = x.strip().lower().replace(",", ".")
        m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*min", s)
        if m:
            return float(m.group(1))
        parts = s.split("/")
        if len(parts) >= 2:
            try:
                return float(re.sub(r"[^\d\.]", "", parts[1]))
            except Exception:
                pass
        try:
            return float(re.sub(r"[^\d\.]", "", s))
        except Exception:
            return 0.0
    return 0.0


# ======================================================
# Capacités véhicules
# ======================================================
def adjust_vehicle_capacities(vehicles):
    def extract_additional_capacity(info):
        if pd.isna(info):
            return 0
        m = re.search(r"(\d+)\s*kg", str(info))
        return int(m.group(1)) if m else 0

    vehicles = vehicles.copy()
    base = pd.to_numeric(vehicles.get("Poids (kg)"), errors="coerce").fillna(0)
    extra_series = vehicles.get("Informations supplémentaires", pd.Series(0, index=vehicles.index)).apply(extract_additional_capacity)
    vehicles["Capacité ajustée (kg)"] = base + extra_series
    return vehicles


# ======================================================
# MAJ feuilles Excel (compactes)
# ======================================================
def update_kilometrage_file(workbook, delivery_data):
    """Rebuild compact 'Kilométrage' replacing dates present in delivery_data."""
    HEADER = ("Date de livraison", "Chauffeur", "Tournée", "Distance (km)")

    def to_ymd(v):
        from datetime import date, datetime as _dt
        if v is None: return None
        if isinstance(v, _dt):  return v.date().isoformat()
        if isinstance(v, date): return v.isoformat()
        s = str(v).strip()
        if " " in s: s = s.split(" ", 1)[0]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try: return _dt.strptime(s, fmt).date().isoformat()
            except: pass
        return s

    new_entries = []
    for e in delivery_data:
        if not e or len(e) < 4: continue
        new_entries.append((to_ymd(e[0]), e[1], e[2], e[3]))
    target_dates = {e[0] for e in new_entries if e[0]}

    kept_rows = []
    if "Kilométrage" in workbook.sheetnames:
        ws_old = workbook["Kilométrage"]
        for r in range(2, (ws_old.max_row or 1) + 1):
            vals = [ws_old.cell(r, c).value for c in range(1, 5)]
            if all(v in (None, "") for v in vals): continue
            d = to_ymd(vals[0])
            if d in target_dates:
                continue
            kept_rows.append((d, vals[1], vals[2], vals[3]))

    tmp_name = "Kilométrage_tmp"
    if tmp_name in workbook.sheetnames:
        workbook.remove(workbook[tmp_name])
    ws_new = workbook.create_sheet(tmp_name)
    for c, v in enumerate(HEADER, start=1):
        ws_new.cell(1, c, v)

    row = 2
    for d, ch, tour, dist in kept_rows + new_entries:
        ws_new.cell(row, 1, d)
        ws_new.cell(row, 2, ch)
        ws_new.cell(row, 3, tour)
        ws_new.cell(row, 4, dist)
        row += 1

    if "Kilométrage" in workbook.sheetnames:
        workbook.remove(workbook["Kilométrage"])
    ws_new.title = "Kilométrage"


def update_distance_parcourue(workbook):
    if not {"Liste", "Kilométrage"}.issubset(workbook.sheetnames):
        return
    ws_list = workbook["Liste"]
    ws_km   = workbook["Kilométrage"]

    totals = defaultdict(float)
    for _, chauffeur, _, dist in ws_km.iter_rows(min_row=2, values_only=True):
        if dist is None: continue
        try: d = float(dist)
        except:
            s = re.sub(r"[^\d\.,]", "", str(dist)).replace(",", ".")
            try: d = float(s)
            except: d = 0.0
        totals[chauffeur] += d

    headers = [cell.value for cell in ws_list[1]]
    if "Distance parcourue (km)" in headers:
        col = headers.index("Distance parcourue (km)") + 1
    else:
        col = len(headers) + 1
        ws_list.cell(1, col, "Distance parcourue (km)")

    for r in range(2, ws_list.max_row + 1):
        nom    = ws_list.cell(r, 1).value
        prenom = ws_list.cell(r, 2).value
        if nom and prenom:
            key = f"{nom} {prenom}"
            ws_list.cell(r, col, totals.get(key, 0.0))
        else:
            ws_list.cell(r, col, None)


# ======================================================
# Distances / aide
# ======================================================
def _compute_route_distance_km(seq_nodes, nodes, km_matrix):
    """km Depot -> seq -> Depot, km_matrix alignée à 'nodes'."""
    if not seq_nodes: return 0.0
    idx = {n: i for i, n in enumerate(nodes)}
    d = km_matrix[idx["FRESH DISTRIB"]][idx[seq_nodes[0]]]
    for a, b in zip(seq_nodes, seq_nodes[1:]):
        d += km_matrix[idx[a]][idx[b]]
    d += km_matrix[idx[seq_nodes[-1]]][idx["FRESH DISTRIB"]]
    return float(d)

def _best_insertion(seq, node, nodes, km_matrix):
    if not seq:
        return [node]
    best_seq, best_cost = None, float("inf")
    for pos in range(len(seq) + 1):
        cand = seq[:pos] + [node] + seq[pos:]
        cost = _compute_route_distance_km(cand, nodes, km_matrix)
        if cost < best_cost:
            best_cost, best_seq = cost, cand
    return best_seq


# ======================================================
# MDS 2D (sans sklearn) + Sweep clustering capacitaire
# ======================================================
def _mds_2d(D):
    """Classical metric MDS en 2D depuis une matrice de distances (numpy array, sym, 0 diag)."""
    n = D.shape[0]
    if n <= 2:
        return np.zeros((n, 2))
    J = np.eye(n) - np.ones((n, n))/n
    D2 = D**2
    B = -0.5 * J @ D2 @ J
    # valeurs propres triées décroissant
    w, V = np.linalg.eigh(B)
    idx = np.argsort(w)[::-1]
    w = w[idx]
    V = V[:, idx]
    # garder les deux positives
    w = np.clip(w[:2], 0, None)
    L = np.diag(np.sqrt(np.where(w > 0, w, 0)))
    X = V[:, :2] @ L
    return X

def _sweep_partition(nodes, km_matrix, w_dem, c_dem, k, cap_w, cap_c, slack=0.05):
    """
    nodes: ['FRESH DISTRIB', c1, c2, ...]
    w_dem, c_dem: dict code -> demand
    cap_w, cap_c: lists per vehicle (length k)
    Retourne: liste de listes de codes clients par véhicule (k items, non vides si possible)
    """
    assert nodes[0] == "FRESH DISTRIB"
    clients = nodes[1:]
    n = len(clients)
    if k <= 0: return [[] for _ in range(0)]
    if n == 0: return [[] for _ in range(k)]
    # Coordonnées MDS
    X = _mds_2d(km_matrix)  # inclut le dépôt
    dep = X[0]
    C = X[1:]
    # angles polaires autour du dépôt
    ang = np.array([math.atan2(c[1]-dep[1], c[0]-dep[0]) for c in C])
    order = np.argsort(ang)
    clients_sorted = [clients[i] for i in order]

    # sweep : on parcourt dans l'ordre angulaire et on remplit véhicule par véhicule
    clusters = [[] for _ in range(k)]
    used_w = [0.0]*k
    used_c = [0.0]*k
    vi = 0
    for cli in clients_sorted:
        demw = float(w_dem.get(cli, 0.0))
        demc = float(c_dem.get(cli, 0.0))
        placed = False
        # essayer de mettre dans veh courant sinon avancer
        for shift in range(k):
            j = (vi + shift) % k
            remw = cap_w[j] - used_w[j]
            remc = cap_c[j] - used_c[j]
            # autorise petit dépassement (slack) pour éviter trous
            if demw <= remw*(1+slack) and demc <= remc*(1+slack):
                clusters[j].append(cli)
                used_w[j] += demw
                used_c[j] += demc
                placed = True
                # si bien rempli, on avance le pointeur
                if used_w[j] > cap_w[j] or used_c[j] > cap_c[j]:
                    vi = (j + 1) % k
                break
        if not placed:
            # forcer dans véhicule le plus libre (min pénalité)
            j = max(range(k), key=lambda t: min(cap_w[t]-used_w[t], cap_c[t]-used_c[t]))
            clusters[j].append(cli)
            used_w[j] += demw
            used_c[j] += demc

    # garantir non-vides si possible
    empties = [i for i, cl in enumerate(clusters) if len(cl) == 0]
    if empties and any(len(cl)>1 for cl in clusters):
        # couper la plus grosse tournée au milieu
        for e in empties:
            donor = max(range(k), key=lambda t: len(clusters[t]))
            if len(clusters[donor]) <= 1: break
            mv = clusters[donor].pop(len(clusters[donor])//2)
            clusters[e].append(mv)
    return clusters


def _cluster_local_repair(clusters, nodes, km_matrix, w_dem, c_dem, cap_w, cap_c, iters=30):
    """
    Petites améliorations inter-clusters : déplacer un client si le total diminue et capacité ok.
    """
    k = len(clusters)
    if k <= 1: return clusters
    idx = {n:i for i,n in enumerate(nodes)}
    def route_cost(cl):
        return _compute_route_distance_km(cl, nodes, km_matrix)

    used_w = [sum(float(w_dem.get(c,0)) for c in clusters[i]) for i in range(k)]
    used_c = [sum(float(c_dem.get(c,0)) for c in clusters[i]) for i in range(k)]
    for _ in range(iters):
        improved = False
        for a in range(k):
            for b in range(k):
                if a==b or not clusters[a]: 
                    continue
                # tester déplacer chaque client de a -> b
                base = route_cost(clusters[a]) + route_cost(clusters[b])
                best = (0.0, None)
                for cli in clusters[a]:
                    dw = float(w_dem.get(cli,0))
                    dc = float(c_dem.get(cli,0))
                    if used_w[b]+dw > cap_w[b] or used_c[b]+dc > cap_c[b]:
                        continue
                    new_a = clusters[a][:]; new_a.remove(cli)
                    new_b = _best_insertion(clusters[b], cli, nodes, km_matrix)
                    new_cost = route_cost(new_a) + route_cost(new_b)
                    delta = new_cost - base
                    if delta < best[0]:
                        best = (delta, cli, new_a, new_b)
                if best[1] is not None and best[0] < -1e-6:
                    # appliquer
                    _, cli, new_a, new_b = best
                    clusters[a] = new_a
                    clusters[b] = new_b
                    used_w[a] -= float(w_dem.get(cli,0)); used_c[a] -= float(c_dem.get(cli,0))
                    used_w[b] += float(w_dem.get(cli,0)); used_c[b] += float(c_dem.get(cli,0))
                    improved = True
                    break
            if improved: break
        if not improved: break
    return clusters


# ======================================================
# Solve TSP (1 véhicule) pour un cluster
# ======================================================
def _solve_single_vehicle_route(cluster_codes, nodes, cost_int, time_limit_s=30):
    """
    cluster_codes: liste de clients (sans dépôt).
    Retourne la séquence optimisée (sans dépôt).
    """
    if not cluster_codes:
        return []
    # construire sous-matrice (dépôt + clients du cluster)
    node_to_idx = {n:i for i,n in enumerate(nodes)}
    sub_nodes = ["FRESH DISTRIB"] + cluster_codes
    map_old_to_new = [node_to_idx[n] for n in sub_nodes]
    sub_cost = cost_int[np.ix_(map_old_to_new, map_old_to_new)]

    mgr = pywrapcp.RoutingIndexManager(len(sub_nodes), 1, 0)
    routing = pywrapcp.RoutingModel(mgr)

    def cb(i, j):
        return int(sub_cost[mgr.IndexToNode(i)][mgr.IndexToNode(j)])
    cb_idx = routing.RegisterTransitCallback(cb)
    routing.SetArcCostEvaluatorOfAllVehicles(cb_idx)

    params = pywrapcp.DefaultRoutingSearchParameters()
    params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
    params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    params.time_limit.seconds = int(max(5, time_limit_s))
    params.log_search = False
    try:
        params.num_search_workers = max(1, os.cpu_count() or 1)
    except Exception:
        pass
    # petits opérateurs efficaces
    try:
        ops = params.local_search_operators
        ops.use_relocate = routing_enums_pb2.BOOL_TRUE
        ops.use_or_opt = routing_enums_pb2.BOOL_TRUE
        ops.use_2opt = routing_enums_pb2.BOOL_TRUE
    except Exception:
        pass

    sol = routing.SolveWithParameters(params)
    if not sol:
        return cluster_codes  # fallback: ordre inchangé

    seq = []
    idx0 = routing.Start(0)
    while not routing.IsEnd(idx0):
        n = mgr.IndexToNode(idx0)
        if n != 0:
            seq.append(sub_nodes[n])
        idx0 = sol.Value(routing.NextVar(idx0))
    return seq


# ======================================================
# Optimisation (Cluster-first, route-second ou direct)
# ======================================================
def run_optimization(
    distance_file,
    orders_file,
    vehicles_file,
    chauffeurs_file,
    critere,
    unavailable_vehicles=None,
    unavailable_chauffeurs=None,

    # cluster-first
    cluster_first: bool = True,
    cluster_capacity_slack: float = 0.05,
    inter_cluster_iters: int = 30,

    # mode direct si cluster_first=False (équilibrage global)
    balance_span: bool = True,
    span_coeff: int = 5,
    time_limit_s: int = 180
):
    # ---------- Lecture sources ----------
    dist_mat_raw = pd.read_excel(distance_file, index_col=0)
    dist_mat_raw.index = dist_mat_raw.index.map(_norm_space)
    dist_mat_raw.columns = dist_mat_raw.columns.map(_norm_space)

    orders       = pd.read_excel(orders_file)
    vehs         = pd.read_excel(vehicles_file, skiprows=1)
    ch_df_full   = pd.read_excel(chauffeurs_file, sheet_name="Liste")

    vehs["Véhicule"] = vehs["Véhicule"].apply(_norm_space)
    for col in ["Nom", "Prénom", "Véhicule affecté", "Statut"]:
        if col not in ch_df_full.columns:
            ch_df_full[col] = ""

    ch_df_full["Nom Complet"] = (ch_df_full["Nom"].astype(str) + " " + ch_df_full["Prénom"].astype(str)).str.strip()
    ch_df_full["norm_name"]   = ch_df_full["Nom Complet"].map(_norm)
    ch_df_full["veh_norm"]    = ch_df_full["Véhicule affecté"].astype(str).map(_norm_space)
    ch_df_full["statut_norm"] = ch_df_full["Statut"].astype(str).str.strip().str.lower().replace({"": "permanent"})

    unv_ch_norm = set(_norm(n) for n in (unavailable_chauffeurs or []))
    unv_vh_set  = set(_norm_space(v) for v in (unavailable_vehicles or []))

    # ---------- Choix des chauffeurs ----------
    perms = ch_df_full[
        (ch_df_full["statut_norm"] == "permanent")
        & (~ch_df_full["norm_name"].isin(unv_ch_norm))
        & (~ch_df_full["veh_norm"].isin(unv_vh_set))
    ][["Nom Complet", "veh_norm"]].rename(columns={"veh_norm": "Véhicule affecté"}).reset_index(drop=True)

    pairs = perms.to_dict(orient="records")

    # temporaires de remplacement (même véhicule) si indisponibles
    if unv_ch_norm:
        need = Counter(
            ch_df_full[ch_df_full["norm_name"].isin(unv_ch_norm) & (~ch_df_full["veh_norm"].isin(unv_vh_set))]["veh_norm"].tolist()
        )
        temps_pool = ch_df_full[
            (ch_df_full["statut_norm"] == "temporaire")
            & (~ch_df_full["norm_name"].isin(unv_ch_norm))
            & (~ch_df_full["veh_norm"].isin(unv_vh_set))
        ][["Nom Complet", "veh_norm"]].rename(columns={"veh_norm": "Véhicule affecté"})

        already = set(p["Nom Complet"] for p in pairs)
        for v, k in need.items():
            cand = temps_pool[temps_pool["Véhicule affecté"] == v]["Nom Complet"].tolist()
            for nm in cand[:k]:
                if nm not in already:
                    pairs.append({"Nom Complet": nm, "Véhicule affecté": v})
                    already.add(nm)

    # filtre véhicules indispo
    pairs = [p for p in pairs if _norm_space(p["Véhicule affecté"]) not in unv_vh_set]
    if not pairs:
        return "Aucun chauffeur disponible après filtrage.", None

    # capacités véhicules
    vehs = adjust_vehicle_capacities(vehs)
    veh_dict = vehs.set_index("Véhicule").to_dict("index")
    cartons_col = next((c for c in vehs.columns if re.search(r"\bcartons?\b", str(c), flags=re.I)), None)

    # ne garder que les paires avec véhicule connu
    pairs = [p for p in pairs if p["Véhicule affecté"] in veh_dict]
    if not pairs:
        return "Aucun véhicule trouvable pour les chauffeurs retenus.", None

    # ordre fixe par véhicule (stabilité)
    pairs = sorted(pairs, key=lambda p: (_norm_space(p["Véhicule affecté"]), _norm(p["Nom Complet"])))
    k = len(pairs)

    # ---------- Demandes ----------
    orders["Code postal"]      = orders["Code postal"].astype(str).str.split(".").str[0]
    orders["Adresse complète"] = orders["Adresse"] + ", " + orders["Code postal"] + ", " + orders["Ville"]
    orders_f = orders[~orders["Libellé"].str.contains("ECHANTILLON RUSTIK", na=False)].copy()

    orders_f["Quantité"] = pd.to_numeric(orders_f["Quantité"], errors="coerce").fillna(0.0)
    orders_f["Cartons"]  = np.where(orders_f["Unité"] == "U",  orders_f["Quantité"]/30.0, 0.0)
    orders_f["Poids"]    = np.where(orders_f["Unité"] == "KG", orders_f["Quantité"],      0.0)
    orders_f["Code client"] = orders_f["Code client"].astype(str).map(_norm_space)

    agg = (
        orders_f.groupby(["Code client","Adresse complète"], dropna=True)
                .agg({"Poids":"sum","Cartons":"sum"})
                .reset_index()
    )
    addr_by_code    = agg.drop_duplicates("Code client").set_index("Code client")["Adresse complète"].to_dict()
    poids_by_code   = dict(zip(agg["Code client"], agg["Poids"]))
    cartons_by_code = dict(zip(agg["Code client"], agg["Cartons"]))

    client_codes = orders_f["Code client"].dropna().astype(str).unique().tolist()
    valid = set(dist_mat_raw.index) & set(dist_mat_raw.columns)
    cust  = [c for c in client_codes if c in valid]
    nodes = ["FRESH DISTRIB"] + cust
    if len(nodes) <= 1:
        return "Aucune commande exploitable.", None

    submat = dist_mat_raw.loc[nodes, nodes]

    km_matrix = submat.applymap(_parse_km_cell).values
    np.fill_diagonal(km_matrix, 0.0)

    if critere == "distance":
        cost_float = km_matrix
    else:
        cost_float = submat.applymap(_parse_min_cell).values
        np.fill_diagonal(cost_float, 0.0)

    if float(np.max(cost_float)) <= 0.0:
        return "Matrice de coûts invalide (toutes valeurs nulles). Vérifie le format km/min.", None

    SCALE = 100
    cost_int = np.rint(cost_float * SCALE).astype(np.int64)
    if cost_int.max() <= 0:
        return "Alerte: coûts nuls après parsing. Vérifie le fichier distances/temps.", None

    # demandes (kg / cartons)
    w_dem = {c: float(poids_by_code.get(c, 0.0)) for c in nodes[1:]}
    c_dem = {c: float(cartons_by_code.get(c, 0.0)) for c in nodes[1:]}

    def cap_w(vid):
        try:    return float(veh_dict[pairs[vid]["Véhicule affecté"]]["Capacité ajustée (kg)"])
        except: return 1e12
    def cap_c(vid):
        if cartons_col is None: return 1e12
        try:    return float(veh_dict[pairs[vid]["Véhicule affecté"]][cartons_col])
        except: return 1e12

    caps_w = [cap_w(v) for v in range(k)]
    caps_c = [cap_c(v) for v in range(k)]

    # ========== CLUSTER-FIRST ==========
    if cluster_first:
        # cas limite: moins de clients que de chauffeurs
        if len(cust) < k:
            k = len(cust)
            pairs = pairs[:k]
            caps_w = caps_w[:k]
            caps_c = caps_c[:k]

        clusters = _sweep_partition(nodes, km_matrix, w_dem, c_dem, k, caps_w, caps_c, slack=cluster_capacity_slack)
        clusters = _cluster_local_repair(clusters, nodes, km_matrix, w_dem, c_dem, caps_w, caps_c, iters=inter_cluster_iters)

        # Solve TSP pour chaque cluster
        per_vehicle_time = max(10, int(time_limit_s // max(1, k)))
        routes = []
        for vid in range(k):
            seq = _solve_single_vehicle_route(clusters[vid], nodes, cost_int, time_limit_s=per_vehicle_time)
            dist_km = _compute_route_distance_km(seq, nodes, km_matrix)
            w_sum = sum(w_dem.get(x,0.0) for x in seq)
            c_sum = sum(c_dem.get(x,0.0) for x in seq)
            routes.append({"vid": vid, "seq": seq, "w": w_sum, "c": c_sum, "dist": dist_km})

    # ========== DIRECT (Multi-véhicules) ==========
    else:
        mgr = pywrapcp.RoutingIndexManager(
            len(cost_int), len(pairs), [0]*len(pairs), [0]*len(pairs)
        )
        routing = pywrapcp.RoutingModel(mgr)

        def dist_cb(i, j): return int(cost_int[mgr.IndexToNode(i)][mgr.IndexToNode(j)])
        cb_idx = routing.RegisterTransitCallback(dist_cb)
        routing.SetArcCostEvaluatorOfAllVehicles(cb_idx)

        # Distance balance
        ub = int(max(1, cost_int.max()) * max(3, len(nodes)))
        routing.AddDimension(cb_idx, 0, ub, True, "Distance")
        dist_dim = routing.GetDimensionOrDie("Distance")
        dist_dim.SetGlobalSpanCostCoefficient(int(span_coeff) if balance_span and span_coeff>0 else 0)

        # Capacités
        wd = [0] + [int(round(w_dem.get(c,0.0))) for c in nodes[1:]]
        cd = [0] + [int(round(c_dem.get(c,0.0))) for c in nodes[1:]]
        w_cb = routing.RegisterUnaryTransitCallback(lambda i: wd[mgr.IndexToNode(i)])
        c_cb = routing.RegisterUnaryTransitCallback(lambda i: cd[mgr.IndexToNode(i)])
        routing.AddDimensionWithVehicleCapacity(w_cb, 0, [int(round(x)) for x in caps_w], True, "Weight")
        routing.AddDimensionWithVehicleCapacity(c_cb, 0, [int(round(x)) for x in caps_c], True, "Cartons")

        params = pywrapcp.DefaultRoutingSearchParameters()
        params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PARALLEL_CHEAPEST_INSERTION
        params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
        params.time_limit.seconds = int(time_limit_s)
        params.log_search = False
        try:
            params.num_search_workers = max(1, os.cpu_count() or 1)
        except Exception:
            pass

        sol = routing.SolveWithParameters(params)
        if not sol:
            return "Aucune solution trouvée.", None

        routes = []
        for vid in range(len(pairs)):
            idx0 = routing.Start(vid)
            seq, w_sum, c_sum = [], 0, 0
            while not routing.IsEnd(idx0):
                n = mgr.IndexToNode(idx0)
                if n != 0:
                    code = nodes[n]
                    seq.append(code)
                    w_sum += wd[n]
                    c_sum += cd[n]
                idx0 = sol.Value(routing.NextVar(idx0))
            dist_km = _compute_route_distance_km(seq, nodes, km_matrix)
            routes.append({"vid": vid, "seq": seq, "w": float(w_sum), "c": float(c_sum), "dist": dist_km})

    # ---------- Export Excel ----------
    wb = Workbook()
    summary = wb.active
    summary.title = "Résumé"
    summary.append(["Tournée", "Lien vers la feuille"])

    result_str = ""
    delivery_data = []

    routes_sorted = sorted(routes, key=lambda r: r["dist"], reverse=True)
    for i, r in enumerate(routes_sorted, start=1):
        chauffeur = pairs[r["vid"]]["Nom Complet"]
        vehicle   = pairs[r["vid"]]["Véhicule affecté"]
        name = sanitize_sheet_name(f"Tournée {i}")
        ws   = wb.create_sheet(name)

        ws.append([f"Chauffeur : {chauffeur}"])
        ws.append([f"Véhicule : {vehicle}"])
        ws.append(["Code client", "Adresse", "Ordre de visite"])

        full_seq = ["FRESH DISTRIB"] + r["seq"] + ["FRESH DISTRIB"]
        for j, code in enumerate(full_seq, start=1):
            adresse = "DEPOT - FRESH DISTRIB" if code == "FRESH DISTRIB" else addr_by_code.get(code, "")
            ws.append([code, adresse, j])

        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=0) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

        summary.append([f"Tournée {i}", f"=HYPERLINK(\"#'{name}'!A1\",\"Aller à {name}\")"])

        result_str += (
            f"Tournée {i} : {chauffeur} via {vehicle}\n"
            f"  Clients  : {' -> '.join(full_seq)}\n"
            f"  Distance : {r['dist']:.0f} km\n"
            f"  Poids    : {r['w']:.1f} kg, Cartons : {r['c']:.1f}\n\n"
        )

        now_str = datetime.now(PARIS_TZ).strftime("%Y-%m-%d") if PARIS_TZ else datetime.now().strftime("%Y-%m-%d")
        delivery_data.append((now_str, chauffeur, f"Tournée {i}", round(r["dist"], 0)))

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    update_kilometrage_file(wb, delivery_data)
    update_distance_parcourue(wb)

    out = BytesIO()
    wb.save(out); out.seek(0)

    total_d = sum(r["dist"] for r in routes)
    total_w = sum(r["w"]    for r in routes)
    total_c = sum(r["c"]    for r in routes)

    empty_left = sum(1 for r in routes if not r["seq"])
    if empty_left:
        result_str += f"⚠️ {empty_left} véhicule(s) sans client (demande insuffisante).\n"

    result_str += f"\nTotal : {int(round(total_d))} km | {total_w:.1f} kg | {total_c:.1f} cartons"
    return result_str, out
